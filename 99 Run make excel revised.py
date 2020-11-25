#!/usr/bin/env python
# coding: utf-8

# ### Load libs

# In[1]:


from openpyxl import Workbook, worksheet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, fills, colors, Color
from openpyxl.formatting.rule import ColorScaleRule
import csv
import pandas as pd
import re


# In[2]:


# set project
project = 'H33 Mars Petcare subsets'


# ### Define helper functions

# In[3]:


def tryToCast(s):
    try:
        s = float(s)
    except:
        try:
            s = int(s)
        except:
            pass
        pass
    return s


def tryToCastInt(s):
    """Tries to cast string to float, if fail returns str"""
    try:
        s = int(s)
    except:
        pass
    return s


def tryToCastFloat(s):
    """Tries to cast string to float, if fail returns str"""
    try:
        if s == '':
            s = 0
        s = float(s)
    except:
        pass
    return s


# In[4]:


# apply border to all cells in merged cell
def style_range(ws, cell_range, border=Border(), fill=None, font=None,
                alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = alignment

    rows = ws[cell_range]
    if font:
        first_cell.font = font

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if fill:
            for c in row:
                c.fill = fill


def getConceptsAndLengths(projectName):
    d = pd.read_excel('Automation V2/' + projectName + '/results/concepts.xlsx')
    return (list(d['Concept']), len(d),list(d['Length']))


def createTitle(ws1, numConcepts, row):
    """Creates concept title merged cell"""
    ws1.merge_cells(start_row=row, start_column=2, end_row=row, end_column=2 + numConcepts - 1)


def PopulateHeaderContents(ws1, concepts, row):
    """Populates contents of header table including column names"""
    for x in range(2, 2 + len(concepts)):
        c = ws1.cell(row=row, column=x)
        c.value = tryToCastInt(concepts[x - 2])
        c.font = Font(b=True, color="000000", size=12, name='Arial Black')
        c.alignment = Alignment(horizontal="center")


def getVibeScores(table):
    numConcepts = len(table[0])-1
    positive = [table[0], table[1], table[2], table[3], table[4], table[7], table[8]]
    negative = [table[5], table[6], table[9], table[10], table[11], table[12], table[13], table[14], table[15],
                table[17]]
    sump = [0 for x in range(numConcepts)]
    sumn = [0 for x in range(numConcepts)]
    for trait in positive:
        for x in range(numConcepts):
            sump[x] += trait[x]

    for trait in negative:
        for x in range(numConcepts):
            sumn[x] += trait[x]
            
    net = []
    for p,n in zip(sump,sumn):
        net.append(p-n)
    vibescores = [((sump[x] / len(positive)) - (sumn[x] / len(negative))) * 100 / 4 for x in range(numConcepts)]
    avgp = [sump[x] / len(positive) for x in range(numConcepts)]
    avgn = [sumn[x] / len(negative) for x in range(numConcepts)]
    net = []
    for p,n in zip(avgp,avgn):
        net.append(p-n)
    return vibescores, sump, sumn, net, avgp, avgn


def AddLabels(ws1, title, concepts, y):
    """Adds row labels for table"""
    x = 1
    for item in range(len(concepts) + 1):
        c = ws1.cell(row=y, column=x)
        if item == 0:
            c.value = title
            c.alignment = Alignment(horizontal="left")
        else:
            c.value = concepts[item-1]
            c.alignment = Alignment(horizontal="center")
        c.font = Font(b=True, color="000000", size=12, name='Arial')
        thin = Side(border_style="thin", color="000000")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        x += 1

    return y + 1


def PopulateTableContentsFloat(ws1, rows, numConcepts, y):
    """Populates contents of header table including column names. Values are cast to floats when applicable"""
    rule = ColorScaleRule(start_type='percentile', start_value=0, start_color='FFF8696B', mid_type='percentile',
                          mid_value=50, mid_color='FFFFEB84', end_type='percentile', end_value=100,
                          end_color='FF63BE7B')
    ws1.conditional_formatting.add("B" + str(y) + ':' + get_column_letter(numConcepts + 1) + str(len(rows) + y - 1),
                                   rule)
    for row in rows:
        x = 1
        for item in row:
            c = ws1.cell(row=y, column=x)
            if x != 1:
                c.value = tryToCastFloat(item)
                c.number_format = '0.00'
            else:
                c.value = tryToCast(item)
            c.font = Font(color="000000", size=12, name='Arial')
            thin = Side(border_style="thin", color="000000")
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if x == 1:
                c.alignment = Alignment(horizontal="left")
            else:
                c.alignment = Alignment(horizontal="center")
            x += 1
        y += 1
    return y + 1


def PopulateTableContentsFloatPlain(ws1, rows, numConcepts, y):
    """Populates contents of header table including column names. Values are cast to floats when applicable"""
    for row in rows:
        x = 1
        for item in row:
            c = ws1.cell(row=y, column=x)
            if x != 1:
                c.value = tryToCastFloat(item)
                c.number_format = '0.00'
            else:
                c.value = tryToCast(item)
            c.font = Font(color="000000", size=12, name='Arial')
            thin = Side(border_style="thin", color="000000")
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if x == 1:
                c.alignment = Alignment(horizontal="left")
            else:
                c.alignment = Alignment(horizontal="center")
            x += 1
        y += 1
    return y + 1


def PopulateTableContentsPercent(ws1, rows, numConcepts, y):
    """Populates contents of header table including column names. Values are cast to floats when applicable"""
    rule = ColorScaleRule(start_type='percentile', start_value=0, start_color='FFF8696B', mid_type='percentile',
                          mid_value=50, mid_color='FFFFEB84', end_type='percentile', end_value=100,
                          end_color='FF63BE7B')
    ws1.conditional_formatting.add("B" + str(y) + ':' + get_column_letter(numConcepts + 1) + str(len(rows) + y - 1),
                                   rule)
    for row in rows:
        x = 1
        for item in row:
            c = ws1.cell(row=y, column=x)
            if x != 1:
                c.value = tryToCastFloat(item)
                c.style = 'Percent'
                c.number_format = '0%'
            else:
                if item == '':
                    c.value = 0
                else:
                    c.value = tryToCast(item)
            c.font = Font(color="000000", size=12, name='Arial')
            thin = Side(border_style="thin", color="000000")
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if x == 1:
                c.alignment = Alignment(horizontal="left")
            else:
                c.alignment = Alignment(horizontal="center")
            x += 1
        y += 1
    return y + 1


def PopulateTableContentsPercentPlain(ws1, rows, y,border=True):
    for row in rows:
        x = 1
        for item in row:
            c = ws1.cell(row=y, column=x)
            if x != 1:
                c.value = tryToCastFloat(item)
                c.style = 'Percent'
                c.number_format = '0%'
            else:
                c.value = tryToCast(item)
            c.font = Font(color="000000", size=12, name='Arial')
            if border:
                thin = Side(border_style="thin", color="000000")
                c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if x == 1:
                c.alignment = Alignment(horizontal="left")
            else:
                c.alignment = Alignment(horizontal="center")
            x += 1
        y += 1
    return y + 1


def PopulateTableContentsString(ws1, rows, y):
    """Populates contents of header table including column names. Values are cast to strings when applicable"""
    for row in rows:
        x = 1
        for item in row:
            c = ws1.cell(row=y, column=x)
            c.value = tryToCastInt(item)
            c.font = Font(color="000000", size=12, name='Arial')
            thin = Side(border_style="thin", color="000000")
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if x == 1:
                c.alignment = Alignment(horizontal="left")
            else:
                c.alignment = Alignment(horizontal="center")
            x += 1
        y += 1
    return y + 1


def PopulateStarTable(ws1, rowsstar, y, numConcepts):
    rule = ColorScaleRule(start_type='percentile', start_value=0, start_color='FFF8696B', mid_type='percentile',
                          mid_value=50, mid_color='FFFFEB84', end_type='percentile', end_value=100,
                          end_color='FF63BE7B')
    ws1.conditional_formatting.add("B" + str(y) + ':' + get_column_letter(numConcepts + 1) + str(2 + y - 2),
                                   rule)
    for row in rowsstar:
        x = 1
        for item in row:
            c = ws1.cell(row=y, column=x)
            if x != 1:
                if row == rowsstar[-1]:
                    c.value = tryToCastFloat(item)
                    c.number_format = '0'
                else:
                    c.value = tryToCastFloat(item)
                    c.number_format = '0.00'
            else:
                c.value = tryToCast(item)
            c.font = Font(color="000000", size=12, name='Arial')
            thin = Side(border_style="thin", color="000000")
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if x == 1:
                c.alignment = Alignment(horizontal="left")
            else:
                c.alignment = Alignment(horizontal="center")
            x += 1
        y += 1
    return y + 1


def CreateHeaderCell(ws1, numConcepts, y, value, count):
    """Create Generic header merged cell"""
    ws1.merge_cells(start_row=y, start_column=2, end_row=y, end_column=2 + numConcepts - 1)

    c = ws1.cell(row=y, column=2)
    c.value = value
    thin = Side(border_style="thin", color="000000")
    b = Border(top=thin, left=thin, right=thin, bottom=thin)
    if numConcepts == 1:
        c.border = b
    else:
        style_range(ws1, str(ws1.merged_cells.ranges[count]), border=b)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.font = Font(b=True, color="000000", size=16, name='Arial')
    y += 2
    return y


def PopulateAttributesTable(ws1, rows, y, numConcepts):
    rule = ColorScaleRule(start_type='percentile', start_value=0, start_color='FFF8696B', mid_type='percentile',
                          mid_value=50, mid_color='FFFFEB84', end_type='percentile', end_value=100,
                          end_color='FF63BE7B')
    ws1.conditional_formatting.add("B" + str(y) + ':' + get_column_letter(numConcepts + 1) + str(len(rows) + y - 2),
                                   rule)
    thin = Side(border_style="thin", color="000000")
    for row in rows[1:]:
        x = 1
        for item in row:
            c = ws1.cell(row=y, column=x)
            if x != 1:
                c.value = tryToCastFloat(item)
                c.style = 'Percent'
                c.number_format = '0%'
            else:
                item = tryToCast(item)
                if type(item) == type(''):

                    c.value = re.findall('^([^:]*)', item.capitalize())[0]
                else:
                    c.value = cleanAttributeRattings(item)
            c.font = Font(color="000000", size=12, name='Arial')
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            if x == 1:
                c.alignment = Alignment(horizontal="left")
            else:
                c.alignment = Alignment(horizontal="center")
            x += 1
        y += 1
    return y + 1


def cleanTrait(s):
    s = s.replace("AI ", "")
    s = s.replace("Express", "Expressive")
    s = s.replace("Distinct", "New News")
    return s


def cleanAttributeRattings(s):
    s.replace(":ATTRIBUTE RATINGS", "")
    return s


# ### Main Function

# In[5]:


def GenerateExcel(projectName, perceptionNames):
    path = 'Automation V2/' + projectName + '/results/'
    path_main = 'Automation V2/' + projectName
    concepts, numConcepts, lengths = getConceptsAndLengths(projectName)
    
    
    wb = Workbook()
    ws1 = wb.active
    ws1.title = 'OVERVIEW'
    
    statsdf = pd.read_excel(path+'stat testing.xlsx')
    
    
    def m(item):
        mapDict = {'HH': 4, 'H': 3, 'M': 2, 'L': 1, '0': 0}
        if item in mapDict.keys():
            return mapDict[item]
        else:
            return item
    
    
    #
    #BEGIN----------------------------------------OVERVIEW SHEET--------------------------------------------BEGIN
    #
    
    #Sets Column Dimensions for Overview Sheet
    for x in range(1 + numConcepts + 1 + 8):
        if x == 0:
            ws1.column_dimensions[get_column_letter(x + 1)].width = 88 + .65
        elif x > 0 and x < numConcepts + 1:
            ws1.column_dimensions[get_column_letter(x + 1)].width = 28 + .65
    
    currentRow = 0
    
    #
    #Overview FIRST HEADER BLOCK
    #
    
    currentRow+=2
    #Headline[A2]
    c = ws1.cell(row=currentRow, column=1)
    c.value = "DETAILED TRAIT SUMMARY"
    c.font = Font(b=False, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="center")

    currentRow+=1
    #Project Name Title[A3]
    c = ws1.cell(row=currentRow, column=1)
    c.value = "Project Name:"
    c.font = Font(b=False, color="000000", size=12, name='Arial')
    c.alignment = Alignment(horizontal="center")

    #Populates Project Name[row3]
    c = ws1.cell(row=currentRow, column=2)
    c.value = projectName
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="center")

    #Creates Project Title Merged Cell [row3]
    createTitle(ws1, numConcepts, currentRow)

    currentRow+=1
    c = ws1.cell(row=currentRow, column=1)
    c.value = "Concept Name:"
    c.font = Font(b=False, color="000000", size=12, name='Arial')
    c.alignment = Alignment(horizontal="center")

    #Concept Names populated in xlsx[row4]
    PopulateHeaderContents(ws1, concepts, currentRow)

    #Change. all numReviews to this source
    currentRow+=1
    c = ws1.cell(row=currentRow, column=1)
    c.value = "Number of reviews:"
    c.font = Font(b=False, color="000000", size=12, name='Arial')
    c.alignment = Alignment(horizontal="center")
    
    #Number of Reviews Actual Count[row5]
    for i,count in enumerate(statsdf.iloc[0][2:]):
        c = ws1.cell(row=currentRow, column=2+i)
        c.value = tryToCastInt(count)
        c.font = Font(b=False, color="000000", size=11, name='Arial')
        c.alignment = Alignment(horizontal="center")

    currentRow+=1
    #Letter association[row6]
    for i,count in enumerate(statsdf.columns[2:]):
        c = ws1.cell(row=currentRow, column=2+i)
        c.value = tryToCastInt(count)
        c.font = Font(b=True, color="000000", size=12, name='Arial')
        c.alignment = Alignment(horizontal="center")

    currentRow+=1
    
    #
    #END: FIRST HEADER BLOCK
    #
    

    #
    #Begin SENTIMENT PERCENTAGES
    #
    
    #Populates Entire Stat Testing.xlsx in place keeping spaces and adding letter markings
    #underneath each data point at A9
    #Also, changes value type to floats for Success Potential and Price Perception statistics
    statsdf.fillna('', inplace=True)
    for i,line in statsdf.iloc[1:].iterrows():
        line = list(line)
        if i%2==1:
            c = ws1.cell(row=currentRow, column=1)
            c.value = line[0]
            c.font = Font(b=False, color="000000", size=11, name='Arial')
            if line[0] in ['SUCCESS POTENTIAL','PRICE PERCEPTION']:
                for i, item in enumerate(line[2:]):
                    c = ws1.cell(row=currentRow, column=2 + i)
                    c.value = tryToCastFloat(item)
                    c.number_format = '0.00'
                    c.font = Font(b=False, color="000000", size=12, name='Arial')
                    c.alignment = Alignment(horizontal="center")
            else:
                for i, item in enumerate(line[2:]):
                    c = ws1.cell(row=currentRow, column=2 + i)
                    c.value = tryToCastFloat(item)
                    c.style = 'Percent'
                    c.number_format = '0%'
                    c.font = Font(b=False, color="000000", size=12, name='Arial')
                    c.alignment = Alignment(horizontal="center")
            currentRow+=1
        else:
            c = ws1.cell(row=currentRow, column=2)
            c.value = line[1]
            c.font = Font(b=True, color="000000", size=11, name='Arial')

            for i, item in enumerate(line[2:]):
                c = ws1.cell(row=currentRow, column=2 + i)
                c.value = item
                c.font = Font(b=False, color="000000", size=12, name='Arial')
                c.alignment = Alignment(horizontal="center")
            currentRow += 2
            
    #
    #End: SENTIMENT PERCENTAGES
    #

    
    #
    #Begin PROJECT SUMMARY-CONCEPT REFERENCE BLOCK
    #
    
    currentRow+=1
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'PROJECT SUMMARY'
    c.font = Font(b=False, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")

    currentRow+=2
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'Title'
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")

    #Project Title Header
    c = ws1.cell(row=currentRow, column=2)
    c.value = projectName
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="center")

    #Merges cells B through numConcepts together
    createTitle(ws1, numConcepts, currentRow)
    
    currentRow+=2
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'Name'
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")

    #Populates concept names
    PopulateHeaderContents(ws1, concepts, currentRow)

    currentRow+=1
    c = ws1.cell(row=currentRow, column=1)
    c.value = '# of reviews'
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")

    #Number of Reviews Actual Count
    for i,count in enumerate(statsdf.iloc[0][2:]):
        c = ws1.cell(row=currentRow, column=2+i)
        c.value = tryToCastInt(count)
        c.font = Font(b=False, color="000000", size=11, name='Arial')
        c.alignment = Alignment(horizontal="center")

    currentRow+=1
    #Concept Reference Cell
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'Concept Reference'
    c.font = Font(b=False, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")

    #Concept Letter Association
    for i in range(numConcepts):
        c = ws1.cell(row=currentRow, column=1 + i + 1)
        c.value = chr(65 + i)
        c.font = Font(b=False, color="000000", size=10, name='Arial')
        c.alignment = Alignment(horizontal="center")

    #
    #End: PROJECT SUMMARY-CONCEPT REFERENCE BLOCK
    #
    
    
    #
    #Begin OVERALL STORY TO NEGATIVE TRAIT VALUES--------------------------TODO------------------------------------
    #
    
    #positive and negative traits 2 xlsx
        
    currentRow+=2
    #Unsure of reason for existence 
    #TODO:
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'OVERALL STORY'
    c.font = Font(b=False, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")

    currentRow+=2
    #Unsure of reason for existence 
    #TODO:
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'ARCHETYPE OF MARKET POTENTIAL'
    c.font = Font(b=False, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")

    currentRow+=1
    #Unsure of reason for existence 
    #TODO:
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'SECONDARY ARCHETYPE'
    c.font = Font(b=False, color="000000", size=10, name='Arial')
    c.alignment = Alignment(horizontal="left")

    currentRow+=2
    #Unsure of reason for existence 
    #TODO:
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'TOP VIBE TRAITS DETECTED'
    c.font = Font(b=False, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    currentRow+=1
    #Unsure of reason for existence 
    #TODO:
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'POSITIVE TRAITS in initial open-ended reaction'
    c.font = Font(b=False, color="000000", size=10, name='Arial')
    c.alignment = Alignment(horizontal="left")
    
    
    
    pTraitNewDf = pd.read_excel(path+'positive traits 2.xlsx') 
    positiveTraitsDf=pTraitNewDf
    
    posTraitFinDf = pd.DataFrame()
    
    for x in range(numConcepts):
        row=0
        for i in range(len(positiveTraitsDf.index)):
            if (pd.isna(positiveTraitsDf['Combo C'+chr(x+49)+' positive'][i]) == False):
                posTraitFinDf.at[row, chr(x+49)] = (positiveTraitsDf['Combo C'+chr(x+49)+' positive'][i])
                row+=1

    
    for x in range(numConcepts):
        for i in range(len(posTraitFinDf.index)):
            c = ws1.cell(row=currentRow+i, column=2+x)
            c.value = posTraitFinDf[chr(x+49)][i]
            c.font = Font(b=False, color="000000", size=10, name='Arial')
            c.alignment = Alignment(horizontal="left")
    
    currentRow+=len(posTraitFinDf.index)
    
    
    #New Vibe Scores List for Calculation
    num_traits=19
    
    newDf = pd.read_excel(path+'trait tallies.xlsx') 
    traitTallyDf = newDf
    
    vibeTally = []
    best_possible_place_holder_list = [4, 4, 4, 4, 4, 0, 0, 0, 4, 4, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    
    
    for i in range(num_traits):
        per_row_list = []
        for x in range(numConcepts):
            valueIdx = concepts[x] + ' value'
            per_row_list.append(traitTallyDf[valueIdx][i])
        per_row_list.append(best_possible_place_holder_list[i])
        vibeTally.append(per_row_list) 
        
    #New Vibe Scores List for Calculation
    
        
    #--------------------------------TODO----------See if we can delete this----------------------------------
    
    traitsNumDf = pd.read_excel(path+'traits.xlsx')
    traitsNumDf = traitsNumDf.applymap(m)
    traitsNumDf['Best Possible'] = [4, 4, 4, 4, 4, 0, 0, 0, 4, 4, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    traitsNumDf.set_index('Trait', inplace = True)
    positiveTraits = ["AI Express", "AI Foresight", "AI Positive", "AI Surprise", "AI Distinct", "AI Intuitive",
                      "AI Excited", "AI New News"]
    negativeTraits = ["AI Negative", "AI Neutral", "AI Fixable", "AI Me-Too", "AI Pointless", "AI Bust", "AI Skeptical",
                      "AI Niche", "AI Overpriced", "AI Nonsense", "AI Unclear", "AI Not For Me"]
    positive = []
    negative = []
    
    positive = []
    negative = []
    for i in range(numConcepts):
        traitListP = []
        traitListN = []
        for trait in traitsNumDf.index[traitsNumDf['AI Prediction C' + str(i + 1)] == 4]:
            if trait.replace(' Emotion', '') in positiveTraits:
                traitListP.append(trait.replace(' Emotion', '') + ' (H+)')
            elif trait.replace(' Emotion', '') in negativeTraits:
                traitListN.append(trait.replace(' Emotion', '') + ' (H+)')
        for trait in traitsNumDf.index[traitsNumDf['AI Prediction C' + str(i + 1)] == 3]:
            if trait.replace(' Emotion', '') in positiveTraits:
                traitListP.append(trait.replace(' Emotion', '') + ' (H)')
            elif trait.replace(' Emotion', '') in negativeTraits:
                traitListN.append(trait.replace(' Emotion', '') + ' (H)')
        for trait in traitsNumDf.index[traitsNumDf['AI Prediction C' + str(i + 1)] == 2]:
            if trait.replace(' Emotion', '') in positiveTraits:
                traitListP.append(trait.replace(' Emotion', '') + ' (M)')
            elif trait.replace(' Emotion', '') in negativeTraits:
                traitListN.append(trait.replace(' Emotion', '') + ' (M)')
        for trait in traitsNumDf.index[traitsNumDf['AI Prediction C' + str(i + 1)] == 1]:
            if trait.replace(' Emotion', '') in positiveTraits:
                traitListP.append(trait.replace(' Emotion', '') + ' (L)')
            elif trait.replace(' Emotion', '') in negativeTraits:
                traitListN.append(trait.replace(' Emotion', '') + ' (L)')
        positive.append(traitListP)
        negative.append(traitListN)

    lengthP = max([len(l) for l in positive])
    lengthN = max([len(l) for l in negative])
    
    for i, traitList in enumerate(positive):
        for j, trait in enumerate(traitList):
            c = ws1.cell(row=currentRow + j, column=1 + i + 1)
            c.value = cleanTrait(trait)
            c.font = Font(b=False, color="000000", size=10, name='Arial')
            c.alignment = Alignment(horizontal="center")

    currentRow += lengthP
    currentRow += 2
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'NEGATIVE TRAITS in initial open-ended reaction'
    c.font = Font(b=False, color="000000", size=10, name='Arial')
    c.alignment = Alignment(horizontal="left")
    
    nTraitNewDf = pd.read_excel(path+'negative traits 2.xlsx') 
    negativeTraitsDf=nTraitNewDf
    
    negTraitFinDf = pd.DataFrame()
    
    for x in range(numConcepts):
        row=0
        for i in range(len(negativeTraitsDf.index)):
            if (pd.isna(negativeTraitsDf['Combo C'+chr(x+49)+' negative'][i]) == False):
                negTraitFinDf.at[row, chr(x+49)] = (negativeTraitsDf['Combo C'+chr(x+49)+' negative'][i])
                row+=1

    
    for x in range(numConcepts):
        for i in range(len(negTraitFinDf.index)):
            c = ws1.cell(row=currentRow+i, column=2+x)
            c.value = negTraitFinDf[chr(x+49)][i]
            c.font = Font(b=False, color="000000", size=10, name='Arial')
            c.alignment = Alignment(horizontal="left")
    
    currentRow+=len(negTraitFinDf.index)

    for i, traitList in enumerate(negative):
        for j, trait in enumerate(traitList):
            c = ws1.cell(row=currentRow + j, column=1 + i + 1)
            c.value = cleanTrait(trait)
            c.font = Font(b=False, color="000000", size=10, name='Arial')
            c.alignment = Alignment(horizontal="center")
    currentRow += lengthN
    
    #--------------------------------TODO----------See if we can delete this----------------------------------
    
    #
    #END: OVERALL STORY TO NEGATIVE TRAIT VALUES-------------------------TODO---------------------------------
    #
     
        
    #
    #Begin OVERALL VIBE SCORE TO PRIMARY SURVEY RATINGS BLOCK
    #
    
    
    #Change list for function
    currentRow += 3
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'OVERALL VIBE SCORE'
    c.font = Font(b=False, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    vibeScores, sump, sumn, net, avgp, avgn = getVibeScores(vibeTally)
    
    #Vibe Score Population
    for x in range(numConcepts):
        c = ws1.cell(row=currentRow, column=1 + x + 1)
        c.value = vibeScores[x]
        c.font = Font(b=False, color="000000", size=12, name='Arial Black')
        c.number_format = '0'
        c.alignment = Alignment(horizontal="center")
    currentRow += 1
    countsDf = pd.read_excel(path+'counts.xlsx')
    
    currentRow += 1
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'EMOTIVE SENTIMENT DETECTION'
    c.font = Font(b=False, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    currentRow += 1
    c = ws1.cell(row=currentRow, column=1)
    c.value = '% POSITIVE '
    c.font = Font(b=False, color="000000", size=10, name='Arial')
    c.alignment = Alignment(horizontal="left")
    index = countsDf.index[countsDf['Trait'].str.contains('Positive')].tolist()[0]
    
    #Emotive Sentiment Positive Percentage Population
    for x in range(numConcepts):
        c = ws1.cell(row=currentRow, column=1 + x + 1)
        c.value = tryToCastFloat(countsDf.iloc[index][1 + x])
        c.style = 'Percent'
        c.font = Font(b=False, color="000000", size=12, name='Arial')
        c.number_format = '0%'
        c.alignment = Alignment(horizontal="center")
    currentRow += 1

    c = ws1.cell(row=currentRow, column=1)
    c.value = '% NEGATIVE '
    c.font = Font(b=False, color="000000", size=10, name='Arial')
    c.alignment = Alignment(horizontal="left")
    index = countsDf.index[countsDf['Trait'].str.contains('Negative')].tolist()[0]
    
    #Emotive Sentiment Negative Percentage Population
    for x in range(numConcepts):
        c = ws1.cell(row=currentRow, column=1 + x + 1)
        c.value = tryToCastFloat(countsDf.iloc[index][1 + x])
        c.style = 'Percent'
        c.font = Font(b=False, color="000000", size=12, name='Arial')
        c.number_format = '0%'
        c.alignment = Alignment(horizontal="center")
    currentRow += 2

    c = ws1.cell(row=currentRow, column=1)
    c.value = 'OVERALL SENTIMENT LEAN'
    c.font = Font(b=False, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")

    #Sentiment Lean Population
    currentRow += 1
    sentDf = pd.read_excel(path+'sent.xlsx')
    currentRow = PopulateTableContentsPercentPlain(ws1, sentDf.values.tolist(), currentRow,border=False)
    starDf = pd.read_excel(path+'star.xlsx')
    
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'PRIMARY SURVEY RATINGS'
    c.font = Font(b=False, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    currentRow += 1

    c = ws1.cell(row=currentRow, column=1)
    c.value = 'SUCCESS POTENTIAL [average, 5-star rating]'
    c.font = Font(b=False, color="000000", size=12, name='Arial')
    c.alignment = Alignment(horizontal="left")

    #Success Potential Score Population
    for x in range(numConcepts):
        c = ws1.cell(row=currentRow, column=1 + x + 1)
        c.value = tryToCastFloat(starDf.iloc[0][x + 1])
        c.font = Font(b=False, color="000000", size=12, name='Arial')
        c.number_format = '0.0'
        c.alignment = Alignment(horizontal="center")
    currentRow += 2
    
    #
    #END: OVERALL VIBE SCORE TO PRIMARY SURVEY RATINGS BLOCK
    #
    
    
    #Row following SUCCESS POTENTIAL is 4.5units in height
    ws1.row_dimensions[currentRow].height = 4.5
    
    
    
    #
    #Personal xlsx sheet not generated/found------------------------------------TODO-------------------------------
    #
    
    try:
        personalDf = pd.read_excel(path+'personal.xlsx')

        currentRow += 1
    
        pd.read_excel(path+'personal.xlsx')
        c = ws1.cell(row=currentRow, column=1)
        c.value = 'PURCHASE INTEREST'
        c.font = Font(b=False, color="000000", size=12, name='Arial')
        c.alignment = Alignment(horizontal="left")

        for x in range(numConcepts):
            c = ws1.cell(row=currentRow, column=1 + x + 1)
            c.value = tryToCastFloat(personalDf.iloc[0][x + 1]) + tryToCastFloat(personalDf.iloc[1][x + 1])
            c.style = 'Percent'
            c.font = Font(b=False, color="000000", size=12, name='Arial')
            c.number_format = '0%'
            c.alignment = Alignment(horizontal="center")
        currentRow += 1

        c = ws1.cell(row=currentRow, column=1)
        c.value = '[% definitely or probably]'
        c.font = Font(b=False, color="000000", size=10, name='Arial')
        c.alignment = Alignment(horizontal="left")
        currentRow += 1
        ws1.row_dimensions[currentRow].height = 4.5
        currentRow += 1
    except:
        print('Can\'t find personal')
     
    #
    #END: Personal xlsx sheet not generated/found--------------------TODO------------------------------------------
    #
    
    
    #
    #BEGIN: ATTRIBUTE RATINGS PERCENTAGES COLOR SCALE BLOCK
    #
    
    currentRow += 1
    
    attributesDf = pd.read_excel(path+'attributes.xlsx')
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'ATTRIBUTE RATINGS'
    c.font = Font(b=False, color="000000", size=10, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    currentRow += 1

    c = ws1.cell(row=currentRow, column=1)
    c.value = '% Strongly or Somewhat Agree'
    c.font = Font(b=False, color="000000", size=10, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    #Concept Names populated in xlsx[row5]
    PopulateHeaderContents(ws1, concepts, currentRow)
    currentRow += 1

    #Populates attribute ratings
    currentRow = PopulateAttributesTable(ws1, attributesDf.values.tolist(), currentRow, numConcepts)
    currentRow += 1
    
    #
    #END: ATTRIBUTE RATINGS PERCENTAGES COLOR SCALE BLOCK
    #
    
    
    #
    #BEGIN: HUNCH TEST AI OUTPUTS------------------------------------TODO--------------------------------------
    #
    
    fillColor = colors.Color(rgb='F2F2F2')
    silverFill = fills.PatternFill(patternType='solid', fgColor=fillColor)
    for x in range(1 + numConcepts):
        ws1.cell(row=currentRow, column=x+1).fill = silverFill
        ws1.cell(row=currentRow+1, column=x+1).fill = silverFill
        ws1.cell(row=currentRow+2, column=x+1).fill = silverFill
   
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'HUNCH TEST AI OUTPUTS'
    c.font = Font(b=True, color="7030a0", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    currentRow += 1
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'PRIMARY ARCHETYPE'
    c.font = Font(b=False, color="000000", size=11, name='Arial')
    c.alignment = Alignment(horizontal="left")

    currentRow+=1
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'SECONDARY ARCHETYPS (if applicable)'
    c.font = Font(b=False, color="000000", size=11, name='Arial')
    c.alignment = Alignment(horizontal="left")
    currentRow += 1 
    
    #No values populate this section yet
    
    #
    #END: HUNCH TEST AI OUTPUTS--------------------------------------------------------------------------
    #
    
    
    #
    #BEGIN: ARCHETYPE MATCH AND LEVELS SCORES
    #

    currentRow += 1
    
    arch1Df = pd.read_excel(path+'arch1.xlsx')
    currentRow = AddLabels(ws1, 'Archetype Match Scores', concepts, currentRow)
    currentRow = PopulateTableContentsFloat(ws1, arch1Df.values.tolist(), numConcepts, currentRow)
    
    arch2Df = pd.read_excel(path+'arch2.xlsx')
    currentRow = AddLabels(ws1, 'Archetype Levels', concepts, currentRow)
    currentRow = PopulateTableContentsString(ws1, arch2Df.values.tolist(), currentRow)
    
    #
    #END: ARCHETYPE MATCH AND LEVELS SCORES
    #
    
    
    #
    #BEGIN: TRAITS AI LEVELS/COUNTS/VALUES BLOCK
    #
    
    currentRow += 1

    currentRow = CreateHeaderCell(ws1, numConcepts, currentRow, 'Traits - AI ORIGINAL', 2)
    
    
    traitsDf = pd.read_excel(path+'traits.xlsx')
    currentRow = AddLabels(ws1, 'Trait Levels', concepts, currentRow)
    currentRow = PopulateTableContentsString(ws1, traitsDf.values.tolist(), currentRow)
    
    
    countsDf = pd.read_excel(path+'counts.xlsx')
    currentRow = AddLabels(ws1, 'Trait Counts', concepts, currentRow)
    currentRow = PopulateTableContentsPercentPlain(ws1, countsDf.values.tolist(), currentRow)
    
    
    traitsNumDf = pd.read_excel(path+'map values.xlsx')
    currentRow = AddLabels(ws1, 'Trait Values', concepts, currentRow)
    currentRow = PopulateTableContentsString(ws1, traitsNumDf.values.tolist(), currentRow)
    
    
    totals = []
    totals.append(['Average Positive']+avgp)
    totals.append(['Average Negative']+avgn)
    totals.append(['Net']+net)

    currentRow = PopulateTableContentsFloatPlain(ws1, totals, numConcepts, currentRow)
    
    #
    #END: TRAITS AI LEVELS/COUNTS/VALUES BLOCK
    #
    
    
    #
    #Begin: VIBE SCORE TO INTEREST LEVEL BLOCK--------------------TODO-----------------------------
    #
    
    #Vibe Score
    currentRow += 1
    c = ws1.cell(row=currentRow, column=1)
    c.value = 'VIBE SCORE - ORIGINAL'
    c.font = Font(b=False, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")

    for x in range(numConcepts):
        c = ws1.cell(row=currentRow, column=1 + x + 1)
        c.value = vibeScores[x]
        c.font = Font(b=False, color="000000", size=12, name='Arial Black')
        c.number_format = '0'
        c.alignment = Alignment(horizontal="center")
    currentRow += 2
    
    #Measures Block
    starDf = pd.read_excel(path+'star.xlsx')
    currentRow = AddLabels(ws1, 'Measures', concepts, currentRow)
    currentRow = PopulateStarTable(ws1, starDf.values.tolist(), currentRow, numConcepts)
    currentRow += 1

    #Interest Level Block
    try:
        currentRow = AddLabels(ws1, 'Interest Level', concepts, currentRow)
        currentRow = PopulateTableContentsPercentPlain(ws1, personalDf.values.tolist(), currentRow)
        currentRow += 1
    except:
        pass
        
    
    genderDf = pd.read_excel(path+'gender.xlsx')
    currentRow = PopulateTableContentsPercentPlain(ws1, genderDf.values.tolist(), currentRow)
    currentRow += 1

    if perceptionNames:
        for name in perceptionNames:
            perceptionDf = pd.read_excel(path + name + "_price.xlsx")
            pt = list(map(list, zip(*perceptionDf.values.tolist())))[1:]

            for x in range(len(concepts)):
                avg = 0
                for y in range(11):
                    if pt[x][y] == '':
                        pt[x][y] = 0
                    avg += tryToCast(pt[x][y]) * y
                c = ws1.cell(row=perceptionRow, column=2 + x)
                c.value = avg
                c.font = Font(b=False, color="000000", size=12, name='Arial')
                c.number_format = '0.0'
                c.alignment = Alignment(horizontal="center")
            perceptionRow += 1
            currentRow = AddLabels(ws1, name, concepts, currentRow)
            currentRow = PopulateTableContentsPercent(ws1, rowsPerception, numConcepts, currentRow)
            currentRow += 1
                   
    #
    #END: VIBE SCORE TO INTEREST LEVEL BLOCK--------------------TODO-----------------------------
    #
    
    #
    #BEGIN: R&D TABLE-------------------------------TODO-----------------------------
    #
        
    try:
        rdDf = pd.read_excel(path + 'R&D.xlsx')
        currentRow = AddLabels(ws1, 'R&D', concepts, currentRow)
        currentRow = PopulateTableContentsPercent(ws1, rdDf.values.tolist(), numConcepts, currentRow)
        currentRow += 1
    except:
        print('Can\'t find R&D')
        
    #
    #END: R&D TABLE------------------------------TODO-----------------------------
    #
        
    #
    #END----------------------------------------OVERVIEW SHEET------------------------------------------------END
    #

        
    #
    #BEGIN----------------------------------- HIGHLIGHTS SHEET ---------------------------------------------BEGIN
    #    
        
    wsh = wb.create_sheet(str('HIGHLIGHTS'))
    
    #Column Width Set
    for x in range(1 + numConcepts + 1 + 8):
        if x == 0:
            wsh.column_dimensions[get_column_letter(x + 1)].width = 112 + .65
        elif x > 0 and x < numConcepts + 1:
            wsh.column_dimensions[get_column_letter(x + 1)].width = 25 + .65
    
    currentRow = 2
    PopulateHeaderContents(wsh, concepts, currentRow)
    
    currentRow+=1
    c = wsh.cell(row=currentRow, column=1)
    c.value = "Sample Size"
    c.font = Font(b=False, color="000000", size=10, name='Arial')
    c.alignment = Alignment(horizontal="left")
    
    for i,count in enumerate(statsdf.iloc[0][2:]):
        c = wsh.cell(row=currentRow, column=2+i)
        c.value = tryToCastInt(count)
        c.font = Font(b=False, color="000000", size=11, name='Arial')
        c.alignment = Alignment(horizontal="center")
    
    currentRow+=1
    c = wsh.cell(row=currentRow, column=1)
    c.value = "REFERENCE FOR STAT TESTING - letters indicate a higher skew (90% confidence level)"
    c.font = Font(b=False, color="000000", size=10, name='Arial')
    c.alignment = Alignment(horizontal="left")
    
    for i,count in enumerate(statsdf.columns[2:]):
        c = wsh.cell(row=currentRow, column=2+i)
        c.value = tryToCastInt(count)
        c.font = Font(b=False, color="000000", size=12, name='Arial')
        c.alignment = Alignment(horizontal="center")

    blackFill = PatternFill(fgColor=Color('404040'), fill_type = "solid")
    grayFill = PatternFill(fgColor=Color('f2f2f2'), fill_type = "solid")
        
    currentRow+=2
    for x in range(1 + numConcepts):
        wsh.cell(row=currentRow, column=x+1).fill = blackFill
        wsh.cell(row=currentRow+1, column=x+1).fill = grayFill
        
    c = wsh.cell(row=currentRow, column=1)
    c.value = "MARKET POTENTIAL ARCHETYPE"
    c.font = Font(b=True, color="ffffff", size=14, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    currentRow+=1
    c = wsh.cell(row=currentRow, column=1)
    c.value = "OVERALL VIBE SCORE(of initial reactions)"
    c.font = Font(b=True, color="000000", size=14, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    for x in range(numConcepts):
        c = wsh.cell(row=currentRow, column=1 + x + 1)
        c.value = vibeScores[x]
        c.font = Font(b=False, color="000000", size=14, name='Arial Black')
        c.number_format = '0'
        c.alignment = Alignment(horizontal="center")
    
    currentRow+=1
    c = wsh.cell(row=currentRow, column=1)
    c.value = "The average score for all positive traits (0-4) minus the average score for all negative traits (0-4) indexed to the perfect market hit score (4). \r\nARROWS indicate scores that fall in the top third of all 21 ideas."
    c.font = Font(b=False, color="000000", size=10, name='Arial')
    c.alignment = Alignment(horizontal="left", wrapText=True)
    
    #To row 10
    currentRow+=2
    for x in range(1 + numConcepts):
        wsh.cell(row=currentRow, column=x+1).fill = grayFill
    
    c = wsh.cell(row=currentRow, column=1)
    c.value = "VIBE PROFILE (most notable traits detected)"
    c.font = Font(b=True, color="000000", size=14, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    #To row 11
    currentRow+=1
    c = wsh.cell(row=currentRow, column=1)
    c.value = "KEY FOR LEVEL DETECTED: \r\nH+ = Very High \r\nH = High \r\nM = Moderate \r\nL = Low"
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left", wrapText=True)
    
    #To row 13
    currentRow+=2
    for x in range(1 + numConcepts):
        wsh.cell(row=currentRow, column=x+1).fill = grayFill
        
    c = wsh.cell(row=currentRow, column=1)
    c.value = "OVERALL TONE"
    c.font = Font(b=True, color="000000", size=14, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    #Row15
    currentRow+=2
    #TODO: Figure parameters for asterisk to make conditional
    c = wsh.cell(row=currentRow, column=1)
    c.value = "% HIGHLY EXPRESSIVE"
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    #Row17
    currentRow+=2
    c = wsh.cell(row=currentRow, column=1)
    c.value = "% HIGHLY EXPRESSIVE (2+)"
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    #Row19
    currentRow+=2
    c = wsh.cell(row=currentRow, column=1)
    c.value = "% POSITIVE"
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    #Row21
    currentRow+=2
    c = wsh.cell(row=currentRow, column=1)
    c.value = "% NEGATIVE"
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    #Row23
    currentRow+=2
    c = wsh.cell(row=currentRow, column=1)
    c.value = "% NEUTRAL"
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    
    currentRow=15
        
    hToneDf=pd.DataFrame()
    hToneDf = hToneDf.append(statsdf.iloc[1:5])
    hToneDf = hToneDf.append(statsdf.iloc[7:9])
    hToneDf = hToneDf.append(statsdf.iloc[13:17])
    
    del hToneDf['Trait']
    del hToneDf['Value']
    hToneDf = hToneDf.reset_index(drop=True)
    
    x=0
    curCol = 2

    for i in range(10):
        curCol = 2
        for x in range(numConcepts):
            if (hToneDf[chr(x+65)][i] == ' '):
                break
            elif (i % 2) == 0:
                c = wsh.cell(row=currentRow, column=curCol)
                c.value = hToneDf[chr(x+65)][i]
                c.style = 'Percent'
                c.number_format = '0%'
                c.font = Font(b=True, size=14, name='Arial Black')
                c.alignment = Alignment(horizontal="center")
            else:
                c = wsh.cell(row=currentRow, column=curCol)
                c.value = hToneDf[chr(x+65)][i]
                c.font = Font(b=False, color="000000", size=10, name='Arial')
                c.alignment = Alignment(horizontal="center")
            curCol+=1
        currentRow+=1
        
    currentRow=13
        
    for x in range(1 + numConcepts):
        wsh.cell(row=currentRow, column=x+1).fill = grayFill
        wsh.cell(row=currentRow+2, column=x+1).fill = grayFill
        wsh.cell(row=currentRow+4, column=x+1).fill = grayFill
        wsh.cell(row=currentRow+6, column=x+1).fill = grayFill
        wsh.cell(row=currentRow+8, column=x+1).fill = grayFill
        wsh.cell(row=currentRow+10, column=x+1).fill = grayFill
    
    
    #Reset so we can use this as indexing
    currentRow=26
    
    for x in range(1 + numConcepts):
        wsh.cell(row=currentRow, column=x+1).fill = blackFill
    
    c = wsh.cell(row=currentRow, column=1)
    c.value = "PRIMARY SURVEY RATINGS "
    c.font = Font(b=True, color="ffffff", size=14, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    currentRow+=2
    c = wsh.cell(row=currentRow, column=1)
    c.value = "SUCCESS POTENTIAL (average, 1-5 Star Scale)"
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    currentRow+=2
    c = wsh.cell(row=currentRow, column=1)
    c.value = "PURCHASE INTEREST (% Definitely or Probably)"
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    currentRow+=2
    c = wsh.cell(row=currentRow, column=1)
    c.value = "PRICE PERCEPTION (average, 0-10: 0 \"Way too high\", 10 \"A Steal\", 5 \"Just Right\")"
    c.font = Font(b=True, color="000000", size=12, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    for x in range(1 + numConcepts):
        wsh.cell(row=currentRow-4, column=x+1).fill = grayFill
        wsh.cell(row=currentRow-2, column=x+1).fill = grayFill
        wsh.cell(row=currentRow, column=x+1).fill = grayFill
    
    
    successPot = pd.DataFrame()
    successPot = successPot.append(statsdf.iloc[41:43])
    del successPot['Trait']
    del successPot['Value']
    successPot = successPot.reset_index(drop=True)
    
    currentRow=28
    curCol = 2

    for x in range(numConcepts):
        c = wsh.cell(row=currentRow, column=curCol+x)
        c.value = successPot[chr(x+65)][0]
        c.number_format = '0.00'
        c.font = Font(b=True, size=14, name='Arial Black')
        c.alignment = Alignment(horizontal="center")
    
    currentRow+=1
    
    for x in range(numConcepts):
        c = wsh.cell(row=currentRow, column=curCol+x)
        c.value = successPot[chr(x+65)][1]
        c.font = Font(b=False, size=10, name='Arial')
        c.alignment = Alignment(horizontal="center")
    
    currentRow=35
    
    for x in range(1 + numConcepts):
        wsh.cell(row=currentRow, column=x+1).fill = blackFill
    
    c = wsh.cell(row=currentRow, column=1)
    c.value = "BELIEF STATEMENT RATINGS (% Strongly or Somewhat Agree)"
    c.font = Font(b=True, color="ffffff", size=14, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    #Row36
    currentRow+=1
    PopulateHeaderContents(wsh, concepts, currentRow)
    
    #Row37
    currentRow+=1
    c = wsh.cell(row=currentRow, column=1)
    c.value = "Sample Size"
    c.font = Font(b=False, color="000000", size=10, name='Arial')
    c.alignment = Alignment(horizontal="left")

    for i,count in enumerate(statsdf.iloc[0][2:]):
        c = wsh.cell(row=currentRow, column=2+i)
        c.value = tryToCastInt(count)
        c.font = Font(b=False, color="000000", size=11, name='Arial')
        c.alignment = Alignment(horizontal="center")
    
    #Row38
    currentRow+=1
    c = wsh.cell(row=currentRow, column=1)
    c.value = "REFERENCE FOR STAT TESTING - letters indicate a higher skew (90% confidence level)"
    c.font = Font(b=False, color="000000", size=10, name='Arial')
    c.alignment = Alignment(horizontal="left")
    
    #Row39
    for i,count in enumerate(statsdf.columns[2:]):
        c = wsh.cell(row=currentRow, column=2+i)
        c.value = tryToCastInt(count)
        c.font = Font(b=False, color="000000", size=12, name='Arial')
        c.alignment = Alignment(horizontal="center")
    
    currentRow+=1
    beliefRatings = pd.DataFrame()
    beliefRatings = beliefRatings.append(statsdf.iloc[43:])
    del beliefRatings['Value']
    beliefRatings = beliefRatings.reset_index(drop=True)
    
    curCol=1
    beliefLen = len(beliefRatings['Trait'])
    
    for i in range(beliefLen):
        if (i%2 == 0):
            c = wsh.cell(row=currentRow, column=curCol)
            c.value = beliefRatings['Trait'][i]
            c.font = Font(b=True, size=12, name='Arial Black')
            c.alignment = Alignment(horizontal="left")
            currentRow+=2
            
    del beliefRatings['Trait']
    
    currentRow=39
    for i in range(beliefLen):
        curCol = 2
        for x in range(numConcepts):
            if (beliefRatings[chr(x+65)][i] == ' '):
                break
            elif (i % 2) == 0:
                c = wsh.cell(row=currentRow, column=curCol)
                c.value = beliefRatings[chr(x+65)][i]
                c.style = 'Percent'
                c.number_format = '0%'
                c.font = Font(b=True, size=12, name='Arial Black')
                c.alignment = Alignment(horizontal="center")
            else:
                c = wsh.cell(row=currentRow, column=curCol)
                c.value = beliefRatings[chr(x+65)][i]
                c.font = Font(b=False, color="000000", size=12, name='Arial')
                c.alignment = Alignment(horizontal="center")
            curCol+=1
        currentRow+=1
    
    currentRow=39
    for x in range(1 + numConcepts):
        for i in range(beliefLen):
            if (i%2==0):
                wsh.cell(row=currentRow + i, column=x+1).fill = grayFill

    currentRow = 40+beliefLen
    
    for x in range(1 + numConcepts):
        wsh.cell(row=currentRow, column=x+1).fill = blackFill
    
    c = wsh.cell(row=currentRow, column=1)
    c.value = "PERCEPTUAL RATINGS (average, 0-10 scale)"
    c.font = Font(b=True, color="ffffff", size=14, name='Arial Black')
    c.alignment = Alignment(horizontal="left")
    
    currentRow+=1
    c = wsh.cell(row=currentRow, column=1)
    c.value = "Sample Size"
    c.font = Font(b=False, color="000000", size=10, name='Arial')
    c.alignment = Alignment(horizontal="left")
    
    PopulateHeaderContents(wsh, concepts, currentRow + 1)
    
    for i,count in enumerate(statsdf.iloc[0][2:]):
        c = wsh.cell(row=currentRow, column=2+i)
        c.value = tryToCastInt(count)
        c.font = Font(b=False, color="000000", size=11, name='Arial')
        c.alignment = Alignment(horizontal="center")
    
    currentRow+=1
    c = wsh.cell(row=currentRow, column=1)
    c.value = "REFERENCE FOR STAT TESTING - letters indicate a higher skew (90% confidence level)"
    c.font = Font(b=False, color="000000", size=10, name='Arial')
    c.alignment = Alignment(horizontal="left")
    
    for i,count in enumerate(statsdf.columns[2:]):
        c = wsh.cell(row=currentRow, column=2+i)
        c.value = tryToCastInt(count)
        c.font = Font(b=False, color="000000", size=12, name='Arial')
        c.alignment = Alignment(horizontal="center")
    
    #
    #END--------------------------------------- HIGHLIGHTS SHEET ---------------------------------------------END
    #
    
    #
    #BEGIN -------------------------------- TRAIT TALLIES SHEET ---------------------------------------------BEGIN
    #
    
    wst = wb.create_sheet(str('TRAIT TALLIES'))  
    
    for x in range(1 + numConcepts + 1 + 8):
        if x == 0:
            wst.column_dimensions[get_column_letter(x + 1)].width = 10 + .65
        elif x == 1:
            wst.column_dimensions[get_column_letter(x + 1)].width = 30 + .65
        else:
            wst.column_dimensions[get_column_letter(x + 1)].width = 25 + .65

    c = wst.cell(row=1, column=1)
    c.value = "Type"
    c.font = Font(b=True, color="000000", size=12, name='Arial')
    c.alignment = Alignment(horizontal="left")
    
    c = wst.cell(row=1, column=2)
    c.value = "Trait"
    c.font = Font(b=True, color="000000", size=12, name='Arial')
    c.alignment = Alignment(horizontal="left")
    
    thin = Side(border_style="thin", color="000000")
    
    #19 is number of rows given the number of traits we have for analysis
    #TODO: Consider combining these two columns
    for x in range(19):
        c = wst.cell(row=x+4, column=1)
        if (x+4 == 20):
            c.value = "NEUTRAL"
            c.font = Font(b=False, color="000000", size=12, name='Arial')
        elif (x+4 == 22):
            c.value = "NA"
            c.font = Font(b=False, color="000000", size=12, name='Arial')
        elif (x+4 == 9 or x+4 == 10) or (x+4 >= 13):
            c.value = "NEG"
            c.font = Font(b=False, color="c90000", size=12, name='Arial')
        else:
            c.value = "POS"
            c.font = Font(b=False, color="4eac5b", size=12, name='Arial')
        c.alignment = Alignment(horizontal="left")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
   

    

    for x in range(19):
        c = wst.cell(row=x+4, column=2)
        c.value = traitTallyDf['Trait'][x]
        c.font = Font(b=False, color="000000", size=12, name='Arial')
        c.alignment = Alignment(horizontal="left")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    numColsTraits = (numConcepts*3)
    
    newDf = pd.read_excel(path+'concepts.xlsx') 
    traitHeadDf=newDf
    
    currentRow=2
    curCol=3
    
#          #Number of Reviews Actual Count[row5]
#     for i,count in enumerate(statsdf.iloc[0][2:]):
#         c = ws1.cell(row=currentRow, column=2+i)
#         c.value = tryToCastInt(count)
#         c.font = Font(b=False, color="000000", size=11, name='Arial')
#         c.alignment = Alignment(horizontal="center")
    
    
    for x in range(numConcepts):
        c = wst.cell(row=currentRow, column=curCol)
        c.value = traitHeadDf['Concept'][x]
        c.font = Font(b=True, color="000000", size=12, name='Arial Black')
        c.alignment = Alignment(horizontal="center")
        
        #pull from statsdf instead of concepts
        c = wst.cell(row=currentRow, column=curCol+1)
        c.value = statsdf[chr(x+65)][0]
        c.font = Font(b=False, color="000000", size=12, name='Arial')
        c.alignment = Alignment(horizontal="center")
        
        c = wst.cell(row=currentRow+1, column=curCol)
        c.value = 'Counts'
        c.font = Font(b=False, color="000000", size=12, name='Arial')
        c.alignment = Alignment(horizontal="center")
        
        c = wst.cell(row=currentRow+1, column=curCol+1)
        c.value = '%'
        c.font = Font(b=False, color="000000", size=12, name='Arial')
        c.alignment = Alignment(horizontal="center")
        
        c = wst.cell(row=currentRow+1, column=curCol+2)
        c.value = 'Score'
        c.font = Font(b=False, color="000000", size=12, name='Arial')
        c.alignment = Alignment(horizontal="center")
        
        curCol+=3
        
    #Filling in data for trait tallies sheet
    currentRow=4
    curCol=3
    for x in range(numConcepts):
        for i in range(19):
            #Format for trait tallies.xlsx
            countIdx= 'count ' + concepts[x]
            percentIdx= 'percent ' + concepts[x]
            valueIdx= concepts[x] + ' value'
        
            c = wst.cell(row=currentRow+i, column=curCol+(x*3))
            c.value = traitTallyDf[countIdx][i]
            c.number_format = '0'
            c.font = Font(b=False, size=12, name='Arial')
            c.alignment = Alignment(horizontal="right")
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            c = wst.cell(row=currentRow+i, column=curCol+(x*3)+1)
            c.value = traitTallyDf[percentIdx][i]
            c.number_format = '0%'
            c.font = Font(b=False, size=12, name='Arial')
            c.alignment = Alignment(horizontal="right")
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            c = wst.cell(row=currentRow+i, column=curCol+(x*3)+2)
            c.value = traitTallyDf[valueIdx][i]
            c.number_format = '0'
            c.font = Font(b=False, size=12, name='Arial')
            c.alignment = Alignment(horizontal="right")
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            
        
        
    #Setting excel cell functions for summary statistics   
    currentRow=24
    curCol=5
    for x in range(numConcepts):
        cL = get_column_letter(curCol+(x*3))
        c = wst.cell(row=currentRow, column=curCol+(x*3))
        c.value = '=AVERAGE('+cL+'4:'+cL+'8, '+cL+'11:'+cL+'12)'
        c.font = Font(b=False, size=12, name='Arial')
        c.number_format='0.00'
        c.alignment = Alignment(horizontal="right")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)   
        
        c = wst.cell(row=currentRow+1, column=curCol+(x*3))
        c.value = '=AVERAGE('+cL+'9:'+cL+'10, '+cL+'13:'+cL+'19, '+cL+'21)'
        c.font = Font(b=False, size=12, name='Arial')
        c.number_format='0.00'
        c.alignment = Alignment(horizontal="right")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)   

        c = wst.cell(row=currentRow+2, column=curCol+(x*3))
        c.value = '='+cL+'24-'+cL+'25'
        c.font = Font(b=False, size=12, name='Arial')
        c.number_format='0.00'
        c.alignment = Alignment(horizontal="right")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)  
        
        c = wst.cell(row=currentRow+3, column=curCol+(x*3))
        c.value = '='+cL+'26*100/4'
        c.font = Font(b=False, size=12, name='Arial')
        c.number_format='0.00'
        c.alignment = Alignment(horizontal="right")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)  
    
    curCol=2
    currentRow=24
    
    for x in range(numConcepts*3):
        for i in range(4):
            c=wst.cell(row=currentRow+i, column=curCol+x)
            c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    greenFill = PatternFill(fgColor=Color('ecf0e0'), fill_type = "solid")
    redFill = PatternFill(fgColor=Color('eedddc'), fill_type = "solid")
    
    c = wst.cell(row=currentRow, column=curCol)
    c.value = 'Average POSITIVE Prevalence'
    c.font = Font(b=True, size=11, name='Arial')
    c.alignment = Alignment(horizontal="left")
    c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    c.fill = greenFill
    
    currentRow+=1
    c = wst.cell(row=currentRow, column=curCol)
    c.value = 'Average NEGATIVE Prevalence'
    c.font = Font(b=True, size=11, name='Arial')
    c.alignment = Alignment(horizontal="left")
    c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    c.fill = redFill
    currentRow+=1
    c = wst.cell(row=currentRow, column=curCol)
    c.value = 'NET'
    c.font = Font(b=True, size=11, name='Arial')
    c.alignment = Alignment(horizontal="left")
    c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    currentRow+=1
    c = wst.cell(row=currentRow, column=curCol)
    c.value = 'VIBE'
    c.font = Font(b=True, size=11, name='Arial')
    c.alignment = Alignment(horizontal="left")
    c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    
    currentRow=30
    curCol=2
    
    c = wst.cell(row=currentRow, column=curCol)
    c.value = 'VIBE SCORE'
    c.font = Font(b=True, size=12, name='Arial')
    c.alignment = Alignment(horizontal="left")
    c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    curCol=3
    for x in range(numConcepts):
        c = wst.cell(row=currentRow-1, column=curCol+x)
        c.value = traitHeadDf['Concept'][x]
        c.font = Font(b=True, color="000000", size=12, name='Arial Black')
        c.alignment = Alignment(horizontal="center")
        
        c = wst.cell(row=currentRow, column=curCol+x)
        c.value = vibeScores[x]
        c.font = Font(b=False, size=12, name='Arial')
        c.number_format='0'
        c.alignment = Alignment(horizontal="right")
        c.border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
    
    cL = get_column_letter(2+(numConcepts))
        
    wst.conditional_formatting.add('C30:'+cL+'30', ColorScaleRule(start_type='percentile', start_value=0, start_color='FFF8696B', 
                                                                  mid_type='percentile', mid_value=50, mid_color='FFFFEB84', 
                                                                  end_type='percentile', end_value=100, end_color='FF63BE7B'))

    
    
    
    #
    #END ----------------------------------- TRAIT TALLIES SHEET --------------------------------------- END
    #
        
    # create new sheets
    survey_folder = 'Automation V2/prepared surveys/'
    replcer_cols = ['AI Express', 'AI Foresight', 'AI Positive', 'AI Surprise', 'AI Excitement', 'AI Negative', 'AI Neutral', 'AI New News', 'AI Intuitive', 'AI Unclear', 'AI Fixable', 'AI Not For Me', 'AI Me-Too', 'AI Pointless', 'AI Bust', 'AI Skeptical', 'AI Niche', 'AI Overpriced', 'AI Nonsense']            
    for x in concepts:
        
        survey_folder_expand = survey_folder + x + '/'
        survey_file = 'post modeling transformer {}_results.xlsx'.format(x)
        survey_load = survey_folder_expand + survey_file
        
        x_rep = x.replace(':', ' ')
        
        init_df = pd.read_excel(survey_load)
        init_cols = list(init_df.columns)
        col_rows = []
        for col in init_cols:
            if col in replcer_cols:
                col = col.replace('AI ', '')
                if col == 'Excitement':
                    col = 'Delight'
                elif col == 'Me-Too':
                    col = 'Blah'  
            col_rows.append(col)

        val_rows = init_df.values.tolist()
        rows = [col_rows] + val_rows
        ws = wb.create_sheet(str(x_rep))
        for y, row in enumerate(rows):
            for x, element in enumerate(row):
                c = ws.cell(row=y + 1, column=x + 1)
                c.value = tryToCast(element)
                c.font = Font(color="000000", size=12, name='Arial')
                
    # adding all surveys
    all_surveys_folder = 'Automation V2/{}/results/'.format(project)
    all_surveys_file = 'ALL SURVEYS.xlsx'
    all_surveys_path = all_surveys_folder + all_surveys_file
    all_srveys = pd.read_excel(all_surveys_path)
    
    init_cols = list(all_srveys.columns)
    val_rows = all_srveys.values.tolist()
    rows = [init_cols] + val_rows
    ws = wb.create_sheet(str('ALL SURVEYS'))
    for y, row in enumerate(rows):
        for x, element in enumerate(row):
            c = ws.cell(row=y + 1, column=x + 1)
            c.value = tryToCast(element)
            c.font = Font(color="000000", size=12, name='Arial')
    
    
    saved = False
    while not saved:
        try:
            print('exporting to {}'.format(path +'{}.xlsx'.format(projectName)))
            wb.save(filename = path +'{}.xlsx'.format(projectName))
            saved = True
        except:
            input("File is still open.\nClose it then press Enter to continue...")
    
    return


# ### Run generate excel

# In[6]:


# generate excel
wb_init = GenerateExcel(project, None)


# In[7]:


print('complete...')


# In[ ]:





# ### End

# In[ ]:




